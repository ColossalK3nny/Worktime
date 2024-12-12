import pandas as pd
from openpyxl import load_workbook

def calculate_work_hours(input_file, sheet_name="Munkaidő összesítő"):
    # Load the original file
    df = pd.read_excel(input_file, engine='openpyxl')
    
    # Convert "Időpont" to datetime for calculations
    df['Időpont'] = pd.to_datetime(df['Időpont'])
    
    # Sort data by employee (Törzsszám) and date-time (Időpont)
    df = df.sort_values(by=["Törzsszám", "Időpont"])
    
    # Separate entries and exits
    entries = df[df["Mozgáskód neve"].str.contains("BE irány")]
    exits = df[df["Mozgáskód neve"].str.contains("KI irány")]
    
    # Merge entries and exits on employee ID (Törzsszám) and date (Dátum)
    summary = pd.merge(
        entries,
        exits,
        on=["Törzsszám", "Dátum"],
        suffixes=("_entry", "_exit")
    )
    
    # Calculate work duration in seconds
    summary['Munkaidő (mp)'] = (summary['Időpont_exit'] - summary['Időpont_entry']).dt.total_seconds()
    
    # Subtract 30 minutes (1800 seconds) for non-Termelés osztály
    summary['Osztály'] = summary['Osztály_entry']  # Copy the osztály for easier referencing
    summary.loc[summary['Osztály'] != 'Termelés', 'Munkaidő (mp)'] -= 1800  # 30 minutes = 1800 seconds
    
    # Convert the total time to hh:mm format
    summary['Munkaidő (óra:perc)'] = summary['Munkaidő (mp)'].apply(lambda x: f"{int(x // 3600)}:{int((x % 3600) // 60):02d}")
    
    # Create summary table
    summary_table = summary[["Kártyatulajdonos_entry", "Dátum", "Munkaidő (óra:perc)", "Osztály"]]
    summary_table.rename(columns={"Kártyatulajdonos_entry": "Kártyatulajdonos"}, inplace=True)
    
    # Load the original workbook
    workbook = load_workbook(input_file)
    
    # Check if the new sheet already exists, if so, remove it
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    
    # Write the new DataFrame to a new sheet
    with pd.ExcelWriter(input_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        summary_table.to_excel(writer, sheet_name=sheet_name, index=False)
    
    print(f"Munkafolyamat kész! Az eredmények az '{sheet_name}' nevű munkalapon találhatók.")

# Használat:
input_file = "worktime.xlsx"  # Itt kell megadni a heti fájl nevét
calculate_work_hours(input_file)